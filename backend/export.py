import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
from typing import List, Dict, Any

def export_excel(invoices: List[Dict[str, Any]], reconciliation_results: List[Dict[str, Any]], output_path: str):
    """
    Export professionally formatted multi-sheet Excel file.
    """
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    currency_format = '₹#,##0.00'
    center_align = Alignment(horizontal="center", vertical="center")
    
    status_colors = {
        "Matched": "C6EFCE", # Green
        "Mismatch": "FFC7CE", # Red
        "Missing in Portal": "FFEB9C", # Amber
        "Missing in Books": "FFEB9C",
        "Pending": "F2F2F2" # Light gray
    }
    
    # ─── Sheet 1: Invoice Register ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Invoice Register"
    
    inv_df = pd.DataFrame(invoices)
    
    if not inv_df.empty:
        # Reorder columns slightly for better view
        cols = ['invoice_number', 'supplier_gstin', 'buyer_gstin', 'vendor_name', 'date', 
                'taxable_value', 'tax_amount', 'total_amount', 'filing_period', 'status']
        inv_df = inv_df[[c for c in cols if c in inv_df.columns]]
        
        for r_idx, row in enumerate(dataframe_to_rows(inv_df, index=False, header=True), 1):
            ws1.append(row)
            
            # Format rows
            if r_idx == 1: # Header
                for cell in ws1[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
            else: # Data
                # Status color coding
                status_idx = cols.index('status') + 1 if 'status' in cols else -1
                if status_idx > 0:
                    status_val = ws1.cell(row=r_idx, column=status_idx).value
                    if status_val in status_colors:
                        fill = PatternFill(start_color=status_colors[status_val], end_color=status_colors[status_val], fill_type="solid")
                        ws1.cell(row=r_idx, column=status_idx).fill = fill
                
                # Currency formatting
                for c in ['taxable_value', 'tax_amount', 'total_amount']:
                    if c in cols:
                        c_idx = cols.index(c) + 1
                        ws1.cell(row=r_idx, column=c_idx).number_format = currency_format
                        
        ws1.freeze_panes = "A2"
        # Auto-fit columns
        for col in ws1.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws1.column_dimensions[column].width = adjusted_width

    # ─── Sheet 2: Reconciliation Report ─────────────────────────────────────
    ws2 = wb.create_sheet(title="Reconciliation Report")
    
    recon_df = pd.DataFrame(reconciliation_results)
    
    if not recon_df.empty:
        # Add Difference column
        if 'tax_amount' in recon_df.columns and 'external_tax' in recon_df.columns:
            recon_df['tax_difference'] = recon_df['tax_amount'].astype(float).fillna(0) - recon_df['external_tax'].astype(float).fillna(0)
        else:
            recon_df['tax_difference'] = 0.0
            
        cols = ['invoice_number', 'gstin', 'taxable_value', 'external_taxable', 'tax_amount', 'external_tax', 'tax_difference', 'status', 'remarks']
        recon_df = recon_df[[c for c in cols if c in recon_df.columns]]
        
        for r_idx, row in enumerate(dataframe_to_rows(recon_df, index=False, header=True), 1):
            ws2.append(row)
            
            if r_idx == 1:
                for cell in ws2[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
            else:
                status_idx = cols.index('status') + 1 if 'status' in cols else -1
                if status_idx > 0:
                    status_val = ws2.cell(row=r_idx, column=status_idx).value
                    if status_val in status_colors:
                        fill = PatternFill(start_color=status_colors[status_val], end_color=status_colors[status_val], fill_type="solid")
                        ws2.cell(row=r_idx, column=status_idx).fill = fill
                        
                for c in ['taxable_value', 'external_taxable', 'tax_amount', 'external_tax', 'tax_difference']:
                    if c in cols:
                        c_idx = cols.index(c) + 1
                        ws2.cell(row=r_idx, column=c_idx).number_format = currency_format

        ws2.freeze_panes = "A2"
        for col in ws2.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws2.column_dimensions[column].width = adjusted_width
            
    # ─── Sheet 3: Summary ───────────────────────────────────────────────────
    ws3 = wb.create_sheet(title="Summary")
    
    summary_data = [
        ["Metric", "Value"],
        ["Total Invoices", len(invoices)],
        ["Total Tax", sum(float(i.get('tax_amount', 0)) for i in invoices)],
    ]
    
    status_counts = {}
    for inv in invoices:
        status_counts[inv.get('status', 'Pending')] = status_counts.get(inv.get('status', 'Pending'), 0) + 1
        
    for k, v in status_counts.items():
        summary_data.append([f"Status: {k}", v])
        
    for r_idx, row in enumerate(summary_data, 1):
        ws3.append(row)
        if r_idx == 1:
            for cell in ws3[1]:
                cell.font = header_font
                cell.fill = header_fill
                
    ws3.cell(row=3, column=2).number_format = currency_format
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 15
    
    wb.save(output_path)
